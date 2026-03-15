from django.shortcuts import render
from django.db.models import Sum
from django.utils import timezone
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from orders.models import Order, OrderItem
from products.models import Product
from decimal import Decimal

@staff_member_required(login_url='login')
def admin_dashboard(request):
    now = timezone.now()
    month_orders = Order.objects.filter(created_at__year=now.year, created_at__month=now.month)
    sales_month = month_orders.aggregate(total=Sum("total"))["total"] or 0
    orders_count = month_orders.count()
    income = sales_month

    top_qs = (
        OrderItem.objects.filter(order__created_at__year=now.year, order__created_at__month=now.month)
        .values("product__name")
        .annotate(qty=Sum("quantity"))
        .order_by("-qty")[:5]
    )
    top_labels = [x["product__name"] for x in top_qs]
    top_values = [int(x["qty"] or 0) for x in top_qs]

    low_inventory = Product.objects.filter(available=True).order_by("stock")
    low_inventory = [p for p in low_inventory if getattr(p, "low_stock", False)][:10]

    ctx = {
        "sales_month": sales_month,
        "orders_count": orders_count,
        "income": income,
        "top_labels": top_labels,
        "top_values": top_values,
        "low_inventory": low_inventory,
        "month_str": now.strftime("%B %Y"),
    }
    return render(request, "reports/admin_dashboard.html", ctx)


@staff_member_required(login_url='login')
def export_monthly_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    now = timezone.now()
    month_orders = Order.objects.filter(created_at__year=now.year, created_at__month=now.month).select_related('customer')
    total_sales = month_orders.aggregate(total=Sum("total"))["total"] or 0
    
    wb = openpyxl.Workbook()
    
    # Hoja 1: Resumen de Pedidos
    ws1 = wb.active
    ws1.title = "Pedidos del Mes"
    
    # Estilo cabecera
    header_fill = PatternFill(start_color="16a34a", end_color="16a34a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ["ID", "Cliente", "WhatsApp", "Fecha", "Total", "Entrega", "Estado"]
    ws1.append(headers)
    for col_num, cell in enumerate(ws1[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
    
    for o in month_orders:
        ws1.append([
            o.id, 
            o.customer.get_full_name() or o.customer.username, 
            o.customer_phone, 
            o.created_at.strftime("%d/%m/%Y"), 
            o.total, 
            o.get_delivery_method_display(), 
            o.get_status_display()
        ])
    
    # Fila de Total
    ws1.append([]) # Espacio
    total_row = ["", "", "", "TOTAL INGRESOS DEL MES:", total_sales, "", ""]
    ws1.append(total_row)
    last_row = ws1.max_row
    ws1.cell(row=last_row, column=4).font = Font(bold=True)
    ws1.cell(row=last_row, column=5).font = Font(bold=True)
    ws1.cell(row=last_row, column=5).number_format = '"$"#,##0.00'
    
    # Hoja 2: Productos Más Vendidos
    ws2 = wb.create_sheet(title="Más Vendidos")
    ws2.append(["Producto", "Cantidad Vendida"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 20
    
    top_products = (
        OrderItem.objects.filter(order__created_at__year=now.year, order__created_at__month=now.month)
        .values("product__name")
        .annotate(qty=Sum("quantity"))
        .order_by("-qty")
    )
    for item in top_products:
        ws2.append([item["product__name"], item["qty"]])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Reporte_Ventas_{now.strftime("%m_%Y")}.xlsx"'
    wb.save(response)
    return response

# Create your views here.
